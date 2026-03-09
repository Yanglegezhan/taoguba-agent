# -*- coding: utf-8 -*-
"""检查Excel文件的颜色"""
from openpyxl import load_workbook

wb = load_workbook('output/连板溢价表_彩色版.xlsx')
ws = wb.active

print("检查跌停单元格颜色:")
found_limit_down = False

for row in range(2, min(100, ws.max_row+1)):
    for col in range(4, min(25, ws.max_column+1)):
        cell = ws.cell(row, col)
        if cell.value and isinstance(cell.value, str) and '%' in cell.value:
            try:
                pct = float(cell.value.replace('%', '').replace('+', ''))
                if pct <= -9.9:
                    bg_color = cell.fill.start_color.rgb if cell.fill.start_color else None
                    font_color = cell.font.color.rgb if cell.font.color else None
                    print(f"行{row}列{col}: {cell.value}")
                    print(f"  背景色: {bg_color}")
                    print(f"  字体色: {font_color}")
                    found_limit_down = True
            except:
                pass

if not found_limit_down:
    print("未找到跌停单元格")

print("\n检查股票名称列（第3列）的背景色:")
for row in range(2, min(20, ws.max_row+1)):
    cell = ws.cell(row, 3)
    if cell.value and not isinstance(cell.value, (int, float)):
        bg_color = cell.fill.start_color.rgb if cell.fill.start_color else None
        print(f"行{row}: {cell.value}, 背景色: {bg_color}")
