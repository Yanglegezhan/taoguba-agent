# -*- coding: utf-8 -*-
"""
检查复合格式单元格（如"金富科技 6板 AI+算力"）是否正确应用板块背景色
"""

from openpyxl import load_workbook

def check_compound_cells():
    """检查复合格式单元格"""
    
    print("=" * 60)
    print("检查复合格式单元格（前7行）")
    print("=" * 60)
    
    # 加载Excel文件
    wb = load_workbook('../output/连板溢价表_彩色版.xlsx')
    ws = wb.active
    
    # 定义板块颜色
    board_colors = {
        'E8E8E8': '主板',
        'FFD6D6': '科创板',
        'D6E4FF': '创业板',
        'FFF4CC': '北交所'
    }
    
    print("\n前7行（应该是复合格式：股票名字 连板高度 题材）:")
    print("-" * 60)
    
    # 检查前7行（行2-8，因为行1是表头）
    for row_idx in range(2, 9):
        print(f"\n第{row_idx-1}行:")
        
        # 检查所有日期列（从D列开始）
        for col_idx in range(4, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value and isinstance(cell.value, str):
                # 检查是否包含"板"字（表示是复合格式）
                if '板' in cell.value:
                    # 检查背景色
                    color = None
                    if cell.fill and cell.fill.start_color:
                        color = cell.fill.start_color.rgb
                        if color and len(color) == 8:
                            color = color[2:]  # 去掉前两位（透明度）
                    
                    board_name = board_colors.get(color, '无板块色')
                    
                    print(f"  列{cell.column_letter}: {cell.value[:30]}... -> {board_name}")
                    
                    if color not in board_colors:
                        print(f"    ⚠️ 未应用板块背景色！颜色={color}")
    
    print("\n" + "=" * 60)
    print("检查完成！")
    print("=" * 60)


if __name__ == "__main__":
    check_compound_cells()
