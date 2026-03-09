# -*- coding: utf-8 -*-
"""
验证板块背景色是否正确应用到所有股票名称单元格
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def verify_board_colors():
    """验证Excel文件中的板块背景色"""
    
    print("=" * 60)
    print("验证板块背景色应用情况")
    print("=" * 60)
    
    # 加载Excel文件
    wb = load_workbook('../output/连板溢价表_彩色版.xlsx')
    ws = wb.active
    
    # 定义板块颜色
    board_colors = {
        'E8E8E8': '主板',
        'FFD6D6': '科创板',
        'D6E4FF': '创业板',
        'FFF4CC': '北交所'
    }
    
    # 统计信息
    total_cells = 0
    stock_name_cells = 0
    board_colored_cells = 0
    
    # 按列统计
    column_stats = {}
    
    print("\n检查所有单元格...")
    
    # 遍历所有单元格（跳过表头）
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            total_cells += 1
            
            # 检查是否有内容
            if cell.value and isinstance(cell.value, str):
                # 检查是否可能是股票名称（包含中文字符）
                if any('\u4e00' <= c <= '\u9fff' for c in cell.value):
                    stock_name_cells += 1
                    
                    # 检查是否有板块背景色
                    if cell.fill and cell.fill.start_color:
                        color = cell.fill.start_color.rgb
                        if color and len(color) == 8:
                            color = color[2:]  # 去掉前两位（透明度）
                        
                        if color in board_colors:
                            board_colored_cells += 1
                            
                            # 记录列统计
                            col_letter = cell.column_letter
                            if col_letter not in column_stats:
                                column_stats[col_letter] = {
                                    'total': 0,
                                    'colored': 0,
                                    'colors': {}
                                }
                            column_stats[col_letter]['total'] += 1
                            column_stats[col_letter]['colored'] += 1
                            
                            board_name = board_colors[color]
                            if board_name not in column_stats[col_letter]['colors']:
                                column_stats[col_letter]['colors'][board_name] = 0
                            column_stats[col_letter]['colors'][board_name] += 1
    
    # 输出统计结果
    print("\n" + "=" * 60)
    print("统计结果")
    print("=" * 60)
    print(f"总单元格数: {total_cells}")
    print(f"包含中文的单元格数（可能是股票名称）: {stock_name_cells}")
    print(f"应用了板块背景色的单元格数: {board_colored_cells}")
    print(f"覆盖率: {board_colored_cells / stock_name_cells * 100:.1f}%")
    
    print("\n" + "=" * 60)
    print("按列统计")
    print("=" * 60)
    
    for col_letter in sorted(column_stats.keys()):
        stats = column_stats[col_letter]
        print(f"\n列 {col_letter}:")
        print(f"  包含中文的单元格: {stats['total']}")
        print(f"  应用板块色的单元格: {stats['colored']}")
        if stats['colors']:
            print(f"  板块分布:")
            for board_name, count in sorted(stats['colors'].items()):
                print(f"    {board_name}: {count}")
    
    print("\n" + "=" * 60)
    print("验证完成！")
    print("=" * 60)
    
    # 检查是否所有股票名称都有板块色
    if board_colored_cells == stock_name_cells:
        print("\n✓ 所有股票名称单元格都已正确应用板块背景色！")
    else:
        print(f"\n✗ 还有 {stock_name_cells - board_colored_cells} 个股票名称单元格未应用板块背景色")
        print("  这可能是因为：")
        print("  1. 单元格内容不是纯股票名称（如包含连板信息、涨跌幅等）")
        print("  2. 股票代码映射中没有该股票")


if __name__ == "__main__":
    verify_board_colors()
