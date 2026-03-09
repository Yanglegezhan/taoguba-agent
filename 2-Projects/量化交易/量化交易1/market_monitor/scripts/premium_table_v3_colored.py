# -*- coding: utf-8 -*-
"""
连板溢价表V3 - 彩色Excel版本
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
sys.path.append('.')
from premium_table_v3 import PremiumTableV3


def parse_cell_value(cell_value):
    """
    解析单元格值，返回类型和数值
    
    Returns:
        tuple: (type, value)
        type: 'board' (连板), 'pct' (涨跌幅), 'text' (文本), 'empty' (空)
        value: 数值或None
    """
    if pd.isna(cell_value) or cell_value == '' or cell_value == '-':
        return 'empty', None
    
    cell_str = str(cell_value)
    
    # 检查是否是连板（包含"板"字）
    if '板' in cell_str:
        # 提取数字
        try:
            board_num = int(cell_str.split('板')[0].replace('(炸)', ''))
            return 'board', board_num
        except:
            return 'text', None
    
    # 检查是否是涨跌幅（包含%）
    if '%' in cell_str:
        try:
            pct = float(cell_str.replace('%', '').replace('+', ''))
            return 'pct', pct
        except:
            return 'text', None
    
    return 'text', None


def get_board_color(board_num):
    """
    根据连板数返回颜色（优化配色）
    
    1-2板: 浅粉色 #FFE4E1
    3-4板: 浅红色 #FFB6C1
    5-6板: 红色 #FF6B6B
    7板及以上: 深红色 #DC143C
    """
    if board_num <= 2:
        return 'FFE4E1'  # 浅粉色
    elif board_num <= 4:
        return 'FFB6C1'  # 浅红色
    elif board_num <= 6:
        return 'FF6B6B'  # 红色
    else:  # 7板及以上
        return 'DC143C'  # 深红色


def get_pct_style(pct):
    """
    根据涨跌幅返回样式（优化配色）
    
    Returns:
        tuple: (bg_color, font_color)
    """
    if pct <= -9.9:  # 跌停
        return ('404040', 'FFFFFF')  # 灰黑底白字（深灰色，醒目但不刺眼）
    elif pct < -5:  # 跌幅大于5%
        return ('66BB6A', 'FFFFFF')  # 绿底白字
    elif pct < 0:  # 跌幅小于5%
        return (None, '388E3C')  # 深绿字
    elif pct == 0:  # 平盘
        return (None, '757575')  # 灰色字
    elif pct <= 5:  # 涨幅小于等于5%
        return (None, 'D32F2F')  # 红色字
    else:  # 涨幅大于5%
        return ('EF5350', 'FFFFFF')  # 红底白字


def get_board_bg_color(stock_code):
    """
    根据股票代码返回板块背景色（淡色）
    
    Args:
        stock_code: 股票代码（字符串）
    
    Returns:
        str: 背景色（十六进制，不含#）
    """
    if not stock_code or pd.isna(stock_code):
        return None
    
    code_str = str(stock_code).zfill(6)
    
    if code_str.startswith('688'):
        return 'FFD6D6'  # 科创板 - 淡红色（更明显）
    elif code_str.startswith('300') or code_str.startswith('301'):
        return 'D6E4FF'  # 创业板 - 淡蓝色（更明显）
    elif code_str.startswith('8') or code_str.startswith('4'):
        return 'FFF4CC'  # 北交所 - 淡黄色（更明显）
    else:
        return 'E8E8E8'  # 主板 - 淡灰色（更明显）


def apply_cell_style(cell, cell_value, stock_name_to_code=None):
    """应用单元格样式
    
    Args:
        cell: 单元格对象
        cell_value: 单元格值
        stock_name_to_code: 股票名称到代码的映射（用于判断是否是股票名称）
    """
    cell_type, value = parse_cell_value(cell_value)
    
    # 设置对齐方式
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    cell.border = thin_border
    
    # 检查是否是股票名称（如果提供了映射字典）
    if stock_name_to_code and isinstance(cell_value, str):
        # 尝试直接匹配
        if cell_value in stock_name_to_code:
            # 这是一个纯股票名称，应用板块背景色
            stock_code = stock_name_to_code[cell_value]
            bg_color = get_board_bg_color(stock_code)
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.font = Font(color='000000')
            return  # 股票名称单元格不需要其他样式
        
        # 尝试从复合字符串中提取股票名称（格式：股票名字 连板高度 题材）
        # 例如："金富科技 6板 AI+算力" -> "金富科技"
        for stock_name in stock_name_to_code.keys():
            if cell_value.startswith(stock_name + ' '):
                # 找到匹配的股票名称，应用板块背景色
                stock_code = stock_name_to_code[stock_name]
                bg_color = get_board_bg_color(stock_code)
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.font = Font(color='000000')
                return  # 股票名称单元格不需要其他样式
    
    if cell_type == 'board':
        # 连板样式
        bg_color = get_board_color(value)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.font = Font(bold=True, color='000000')
        
    elif cell_type == 'pct':
        # 涨跌幅样式
        bg_color, font_color = get_pct_style(value)
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        if font_color:
            cell.font = Font(bold=True, color=font_color)
    
    else:
        # 普通文本
        cell.font = Font(color='000000')


def generate_colored_excel(df, output_file, board_df=None):
    """
    生成带颜色的Excel文件
    
    Args:
        df: DataFrame（溢价表）
        output_file: 输出文件路径
        board_df: 连板分析数据（用于获取股票代码）
    """
    print("\n生成彩色Excel文件...")
    
    # 创建股票名称到股票代码的映射（只映射唯一的股票）
    stock_name_to_code = {}
    if board_df is not None:
        print("  创建股票名称映射...")
        # 只保留每个股票名称的第一条记录
        unique_stocks = board_df.drop_duplicates(subset=['股票名称'], keep='first')
        for _, row in unique_stocks.iterrows():
            stock_name_to_code[row['股票名称']] = str(row['股票代码']).zfill(6)
        print(f"  映射了 {len(stock_name_to_code)} 只股票")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "连板溢价表"
    
    print("  写入数据...")
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 表头样式
            if r_idx == 1:
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 所有单元格应用样式（传入stock_name_to_code以识别股票名称）
                apply_cell_style(cell, value, stock_name_to_code)
    
    print("  设置列宽...")
    # 设置列宽
    ws.column_dimensions['A'].width = 6   # 序号
    ws.column_dimensions['B'].width = 25  # 概念
    ws.column_dimensions['C'].width = 12  # 股票名称
    
    # 日期列宽度
    for col_idx in range(4, len(df.columns) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 12
    
    # 冻结前3列和第1行
    ws.freeze_panes = 'D2'
    
    print("  保存文件...")
    # 保存
    wb.save(output_file)
    print(f"已保存到: {output_file}")


def main(auto_mode=False):
    """主函数
    
    Args:
        auto_mode: 自动模式，不询问日期，直接使用最新日期
    """
    print("=" * 60)
    print("生成连板溢价表V3 - 彩色Excel版")
    print("=" * 60)
    
    # 1. 加载数据
    print("\n1. 加载连板分析数据...")
    board_df = pd.read_csv('output/board_analysis.csv', encoding='utf-8-sig')
    print(f"数据形状: {board_df.shape}")
    
    # 显示数据日期范围
    board_df['日期'] = pd.to_datetime(board_df['日期'])
    latest_date = board_df['日期'].max()
    earliest_date = board_df['日期'].min()
    print(f"数据日期范围: {earliest_date.date()} 至 {latest_date.date()}")
    
    # 2. 加载概念信息
    print("\n2. 获取概念信息...")
    concept_dict = {}
    
    # 首先加载手动维护的概念
    try:
        from manual_concepts import get_all_concepts
        manual_concepts = get_all_concepts()
        concept_dict.update(manual_concepts)
        print(f"  已加载 {len(manual_concepts)} 只股票的手动概念")
    except:
        print("  未找到手动概念文件")
    
    # 检查是否有缓存的概念数据
    concept_cache_file = 'output/concept_cache.csv'
    if os.path.exists(concept_cache_file):
        print("  从缓存加载概念信息...")
        try:
            # 指定股票代码列为字符串类型
            concept_cache = pd.read_csv(concept_cache_file, encoding='utf-8-sig', dtype={'股票代码': str})
            # 确保股票代码是6位字符串
            concept_cache['股票代码'] = concept_cache['股票代码'].str.zfill(6)
            cached_concepts = dict(zip(concept_cache['股票代码'], concept_cache['概念']))
            # 手动概念优先级更高，不覆盖
            for code, concept in cached_concepts.items():
                if code not in concept_dict:
                    concept_dict[code] = concept
            print(f"  已加载 {len(cached_concepts)} 只股票的缓存概念")
        except Exception as e:
            print(f"  缓存加载失败: {e}")
    
    print(f"  总计: {len(concept_dict)} 只股票有概念信息")
    
    # 3. 询问用户指定日期（除非是自动模式）
    end_date = None
    if not auto_mode:
        print("\n" + "=" * 60)
        print("请输入结束日期（格式：YYYY-MM-DD）")
        print(f"直接回车使用最新日期: {latest_date.date()}")
        print("=" * 60)
        
        user_input = input("结束日期: ").strip()
        
        if user_input:
            try:
                # 验证日期格式
                end_date = pd.to_datetime(user_input)
                print(f"✓ 使用指定日期: {end_date.date()}")
            except:
                print(f"✗ 日期格式错误，使用最新日期: {latest_date.date()}")
                end_date = None
        else:
            print(f"✓ 使用最新日期: {latest_date.date()}")
    else:
        print(f"\n自动模式：使用最新日期 {latest_date.date()}")
    
    # 4. 生成表格
    print("\n3. 生成连板溢价表...")
    generator = PremiumTableV3(board_df, concept_dict)
    premium_table = generator.generate_table(end_date=end_date, days=20)
    
    # 5. 保存CSV版本
    csv_file = 'output/连板溢价表_最终版.csv'
    premium_table.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"\n已保存CSV版本: {csv_file}")
    
    # 6. 生成彩色Excel版本
    excel_file = 'output/连板溢价表_彩色版.xlsx'
    generate_colored_excel(premium_table, excel_file, board_df)
    print("\n表格预览（前10行，前6列）:")
    print(premium_table.iloc[:10, :6])
    
    print("\n" + "=" * 60)
    print("完成！")
    print("=" * 60)
    
    print("\n颜色说明：")
    print("【连板颜色】（优化配色）")
    print("  1-2板: 浅粉色 (#FFE4E1)")
    print("  3-4板: 浅红色 (#FFB6C1)")
    print("  5-6板: 红色 (#FF6B6B)")
    print("  7板及以上: 深红色 (#DC143C)")
    print("\n【涨跌幅颜色】（优化配色）")
    print("  跌停 (≤-9.9%): 黑底白字 (#000000) ⚫ 醒目")
    print("  大跌 (-9.9%~-5%): 绿底白字 (#66BB6A)")
    print("  小跌 (-5%~0%): 深绿字 (#388E3C)")
    print("  平盘 (=0%): 灰色字 (#757575)")
    print("  小涨 (0%~5%): 红色字 (#D32F2F)")
    print("  大涨 (>5%): 红底白字 (#EF5350)")
    print("\n【股票名称背景色】（板块区分）")
    print("  主板: 淡灰色 (#E8E8E8)")
    print("  科创板(688): 淡红色 (#FFD6D6)")
    print("  创业板(300/301): 淡蓝色 (#D6E4FF)")
    print("  北交所(8/4开头): 淡黄色 (#FFF4CC)")
    print("\n【前7行格式】")
    print("  格式: 股票名字 连板高度 题材")
    print("  示例: 金富科技 6板 AI+算力")


if __name__ == "__main__":
    main()
