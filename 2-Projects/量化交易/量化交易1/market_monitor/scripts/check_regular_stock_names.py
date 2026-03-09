# -*- coding: utf-8 -*-
"""
检查普通股票名称单元格（第8行起）是否正确应用板块背景色
"""

from openpyxl import load_workbook
import pandas as pd

def check_regular_stock_names():
    """检查普通股票名称单元格"""
    
    print("=" * 60)
    print("检查普通股票名称单元格（第8行起）")
    print("=" * 60)
    
    # 加载Excel文件
    wb = load_workbook('../output/连板溢价表_彩色版.xlsx')
    ws = wb.active
    
    # 加载连板分析数据以获取股票名称列表
    board_df = pd.read_csv('../output/board_analysis.csv', encoding='utf-8-sig')
    stock_names = set(board_df['股票名称'].unique())
    
    # 定义板块颜色
    board_colors = {
        'E8E8E8': '主板',
        'FFD6D6': '科创板',
        'D6E4FF': '创业板',
        'FFF4CC': '北交所'
    }
    
    print(f"\n已知股票名称数量: {len(stock_names)}")
    print("\n检查第8-20行的股票名称单元格...")
    print("-" * 60)
    
    stock_name_cells_found = 0
    stock_name_cells_with_color = 0
    
    # 检查第8-20行
    for row_idx in range(9, 21):  # 行9-20（Excel中的第9-20行）
        # 检查所有日期列（从D列开始）
        for col_idx in range(4, min(10, ws.max_column + 1)):  # 只检查前几列
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value and isinstance(cell.value, str):
                # 检查是否是已知的股票名称
                if cell.value in stock_names:
                    stock_name_cells_found += 1
                    
                    # 检查背景色
                    color = None
                    if cell.fill and cell.fill.start_color:
                        color = cell.fill.start_color.rgb
                        if color and len(color) == 8:
                            color = color[2:]  # 去掉前两位（透明度）
                    
                    if color in board_colors:
                        stock_name_cells_with_color += 1
                        board_name = board_colors[color]
                        print(f"✓ 行{row_idx} 列{cell.column_letter}: {cell.value} -> {board_name}")
                    else:
                        print(f"✗ 行{row_idx} 列{cell.column_letter}: {cell.value} -> 无板块色 (color={color})")
    
    print("\n" + "=" * 60)
    print("统计结果")
    print("=" * 60)
    print(f"找到的股票名称单元格: {stock_name_cells_found}")
    print(f"应用了板块背景色的: {stock_name_cells_with_color}")
    if stock_name_cells_found > 0:
        print(f"覆盖率: {stock_name_cells_with_color / stock_name_cells_found * 100:.1f}%")
    
    if stock_name_cells_with_color == stock_name_cells_found:
        print("\n✓ 所有股票名称单元格都已正确应用板块背景色！")
    else:
        print(f"\n✗ 还有 {stock_name_cells_found - stock_name_cells_with_color} 个股票名称单元格未应用板块背景色")


if __name__ == "__main__":
    check_regular_stock_names()
