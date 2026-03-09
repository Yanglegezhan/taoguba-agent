# -*- coding: utf-8 -*-
"""检查股票名称列的颜色"""
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('output/连板溢价表_彩色版.xlsx')
ws = wb.active

# 加载board_df获取股票代码
board_df = pd.read_csv('output/board_analysis.csv', encoding='utf-8-sig')
stock_name_to_code = {}
for _, row in board_df.drop_duplicates(subset=['股票名称']).iterrows():
    stock_name_to_code[row['股票名称']] = str(row['股票代码']).zfill(6)

print("股票名称列背景色详情（第8行起）:")
print("-" * 80)
print(f"{'行号':<6} {'股票名称':<12} {'股票代码':<10} {'背景色':<12} {'板块'}")
print("-" * 80)

for row in range(8, min(50, ws.max_row+1)):
    cell = ws.cell(row, 3)
    if cell.value and not pd.isna(cell.value):
        stock_name = cell.value
        stock_code = stock_name_to_code.get(stock_name, '未知')
        bg = cell.fill.start_color.rgb if cell.fill.start_color else 'None'
        
        # 判断板块
        if stock_code.startswith('688'):
            board = '科创板'
        elif stock_code.startswith('300') or stock_code.startswith('301'):
            board = '创业板'
        elif stock_code.startswith('8') or stock_code.startswith('4'):
            board = '北交所'
        else:
            board = '主板'
        
        print(f"{row:<6} {stock_name:<12} {stock_code:<10} {bg:<12} {board}")
